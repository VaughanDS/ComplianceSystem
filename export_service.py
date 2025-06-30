# services/export_service.py
"""
Export service for generating reports and data exports
Supports Excel, PDF, and CSV formats with advanced formatting
"""
import logger
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple, Any, Union
import json
import csv
from io import BytesIO, StringIO

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel export limited to basic pandas functionality")

# ReportLab imports for PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from config.settings import get_config
from core.models import Task, TeamMember, LegislationReference
from core.constants import TaskStatus, Priority
from utils.logger import get_logger

logger = get_logger(__name__)


class ReportGenerator:
    """Base class for report generation"""

    def __init__(self):
        self.config = get_config()
        if REPORTLAB_AVAILABLE:
            self.styles = getSampleStyleSheet()
            self._setup_custom_styles()
        else:
            self.styles = None

    def _setup_custom_styles(self):
        """Setup custom PDF styles"""
        if not REPORTLAB_AVAILABLE:
            return

        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30
        ))

        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='Subtitle',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#444444'),
            spaceAfter=20
        ))

        # Footer style
        self.styles.add(ParagraphStyle(
            name='Footer',
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        ))


class ExportService:
    """Handles all data export operations"""

    def __init__(self):
        self.config = get_config()
        self.report_generator = ReportGenerator()
        self.export_dir = self.config.base_path / "Exports"
        self.export_dir.mkdir(exist_ok=True)

    def export_tasks_to_excel(self, tasks: List[Task],
                             filename: Optional[str] = None,
                             include_charts: bool = True) -> str:
        """Export tasks to Excel with formatting and charts"""
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Tasks_Export_{timestamp}.xlsx"

            filepath = self.export_dir / filename

            # Create DataFrame
            df = self._tasks_to_dataframe(tasks)

            if OPENPYXL_AVAILABLE and include_charts:
                # Create Excel file with openpyxl
                wb = Workbook()

                # Add data sheet
                ws_data = wb.active
                ws_data.title = "Tasks"
                self._add_dataframe_to_worksheet(ws_data, df)

                # Add summary sheet
                ws_summary = wb.create_sheet("Summary")
                self._add_task_summary(ws_summary, tasks)

                # Add charts sheet
                ws_charts = wb.create_sheet("Charts")
                self._add_task_charts(ws_charts, tasks)

                # Save workbook
                wb.save(filepath)
            else:
                # Fallback to pandas export
                with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
                    df.to_excel(writer, sheet_name='Tasks', index=False)

                    # Add summary sheet
                    summary_df = self._create_summary_dataframe(tasks)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)

            logger.info(f"Exported {len(tasks)} tasks to {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting tasks to Excel: {e}")
            raise

    def export_tasks_to_csv(self, tasks: List[Task],
                           filename: Optional[str] = None) -> str:
        """Export tasks to CSV"""
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Tasks_Export_{timestamp}.csv"

            filepath = self.export_dir / filename

            # Create DataFrame and export
            df = self._tasks_to_dataframe(tasks)
            df.to_csv(filepath, index=False)

            logger.info(f"Exported {len(tasks)} tasks to {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting tasks to CSV: {e}")
            raise

    def export_tasks_to_pdf(self, tasks: List[Task],
                           filename: Optional[str] = None,
                           title: str = "Task Report") -> Optional[str]:
        """Export tasks to PDF"""
        if not REPORTLAB_AVAILABLE:
            logger.error("ReportLab not available - cannot export to PDF")
            return None

        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Tasks_Report_{timestamp}.pdf"

            filepath = self.export_dir / filename

            # Create PDF
            doc = SimpleDocTemplate(
                str(filepath),
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18,
            )

            # Build content
            story = []

            # Add title
            story.append(Paragraph(title, self.report_generator.styles['CustomTitle']))
            story.append(Spacer(1, 0.2 * inch))

            # Add metadata
            story.append(Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                self.report_generator.styles['Normal']
            ))
            story.append(Paragraph(
                f"Total Tasks: {len(tasks)}",
                self.report_generator.styles['Normal']
            ))
            story.append(Spacer(1, 0.3 * inch))

            # Add summary
            story.extend(self._create_pdf_summary(tasks))
            story.append(PageBreak())

            # Add task table
            story.append(Paragraph("Task Details", self.report_generator.styles['Heading1']))
            story.append(Spacer(1, 0.2 * inch))

            task_table = self._create_pdf_task_table(tasks)
            story.append(task_table)

            # Build PDF
            doc.build(story)

            logger.info(f"Exported {len(tasks)} tasks to PDF: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting tasks to PDF: {e}")
            return None

    def export_team_to_excel(self, team_members: List[TeamMember],
                            filename: Optional[str] = None) -> str:
        """Export team members to Excel"""
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Team_Export_{timestamp}.xlsx"

            filepath = self.export_dir / filename

            # Create DataFrame
            df = self._team_to_dataframe(team_members)

            if OPENPYXL_AVAILABLE:
                # Create Excel file with formatting
                wb = Workbook()
                ws = wb.active
                ws.title = "Team Members"

                self._add_dataframe_to_worksheet(ws, df)

                # Add department summary
                ws_dept = wb.create_sheet("Department Summary")
                self._add_department_summary(ws_dept, team_members)

                wb.save(filepath)
            else:
                # Fallback to pandas
                df.to_excel(filepath, index=False)

            logger.info(f"Exported {len(team_members)} team members to {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting team to Excel: {e}")
            raise

    def export_legislation_to_excel(self, legislation: List[LegislationReference],
                                   filename: Optional[str] = None) -> str:
        """Export legislation references to Excel"""
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Legislation_Export_{timestamp}.xlsx"

            filepath = self.export_dir / filename

            # Create DataFrame
            df = self._legislation_to_dataframe(legislation)

            if OPENPYXL_AVAILABLE:
                # Create Excel file with formatting
                wb = Workbook()
                ws = wb.active
                ws.title = "Legislation"

                self._add_dataframe_to_worksheet(ws, df)

                # Add category summary
                ws_cat = wb.create_sheet("Category Summary")
                self._add_category_summary(ws_cat, legislation)

                wb.save(filepath)
            else:
                # Fallback to pandas
                df.to_excel(filepath, index=False)

            logger.info(f"Exported {len(legislation)} legislation references to {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting legislation to Excel: {e}")
            raise

    def generate_compliance_report(self, report_data: Dict[str, Any],
                                  format: str = 'excel',
                                  filename: Optional[str] = None) -> str:
        """Generate comprehensive compliance report"""
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Compliance_Report_{timestamp}.{format}"

            filepath = self.export_dir / filename

            if format == 'excel':
                return self._generate_excel_compliance_report(report_data, filepath)
            elif format == 'pdf':
                return self._generate_pdf_compliance_report(report_data, filepath)
            elif format == 'json':
                return self._generate_json_compliance_report(report_data, filepath)
            else:
                raise ValueError(f"Unsupported format: {format}")

        except Exception as e:
            logger.error(f"Error generating compliance report: {e}")
            raise

    def _tasks_to_dataframe(self, tasks: List[Task]) -> pd.DataFrame:
        """Convert tasks to DataFrame"""
        data = []
        for task in tasks:
            data.append({
                'Key': task.key,
                'Title': task.title,
                'Status': task.status,
                'Priority': task.priority,
                'Compliance Area': task.compliance_area,
                'Subcategory': task.subcategory,
                'Task Setter': task.task_setter,
                'Allocated To': ', '.join(task.allocated_to),
                'Manager': task.manager,
                'Created Date': task.created_date,
                'Target Date': task.target_date,
                'Completion Date': task.completion_date,
                'Description': task.description,
                'Days Open': task.days_open,
                'Is Overdue': 'Yes' if task.is_overdue() else 'No'
            })

        return pd.DataFrame(data)

    def _team_to_dataframe(self, team_members: List[TeamMember]) -> pd.DataFrame:
        """Convert team members to DataFrame"""
        data = []
        for member in team_members:
            data.append({
                'Name': member.name,
                'Email': member.email,
                'Department': member.department,
                'Role': member.role,
                'Manager': member.manager,
                'Active': 'Yes' if member.active else 'No',
                'Date Added': member.date_added,
                'Last Login': member.last_login
            })

        return pd.DataFrame(data)

    def _legislation_to_dataframe(self, legislation: List[LegislationReference]) -> pd.DataFrame:
        """Convert legislation references to DataFrame"""
        data = []
        for leg in legislation:
            data.append({
                'Code': leg.code,
                'Full Name': leg.full_name,
                'Category': leg.category,
                'Jurisdiction': leg.jurisdiction,
                'Effective Date': leg.effective_date,
                'Last Updated': leg.last_updated,
                'Summary': leg.summary,
                'Key Requirements': ', '.join(leg.key_requirements[:3]) + '...' if len(leg.key_requirements) > 3 else ', '.join(leg.key_requirements),
                'Penalties': ', '.join(leg.penalties[:2]) + '...' if len(leg.penalties) > 2 else ', '.join(leg.penalties),
                'Review Frequency': leg.review_frequency,
                'Owner': leg.owner,
                'Applicable Areas': ', '.join(leg.applicable_areas)
            })

        return pd.DataFrame(data)

    def _add_dataframe_to_worksheet(self, ws: Worksheet, df: pd.DataFrame):
        """Add DataFrame to worksheet with formatting"""
        if not OPENPYXL_AVAILABLE:
            return

        # Add headers
        for col_num, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Alternate row coloring
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                               min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

    def _add_task_summary(self, ws: Worksheet, tasks: List[Task]):
        """Add task summary to worksheet"""
        if not OPENPYXL_AVAILABLE:
            return

        # Title
        ws['A1'] = 'Task Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        # Summary data
        summary_data = self._calculate_task_summary(tasks)

        row = 3
        for key, value in summary_data.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Format cells
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2):
            row[0].font = Font(bold=True)
            row[1].alignment = Alignment(horizontal='right')

    def _add_task_charts(self, ws: Worksheet, tasks: List[Task]):
        """Add charts to worksheet"""
        if not OPENPYXL_AVAILABLE:
            return

        # Prepare data for charts
        status_counts = {}
        priority_counts = {}

        for task in tasks:
            status_counts[task.status] = status_counts.get(task.status, 0) + 1
            priority_counts[task.priority] = priority_counts.get(task.priority, 0) + 1

        # Add status chart data
        ws['A1'] = 'Status'
        ws['B1'] = 'Count'
        row = 2
        for status, count in status_counts.items():
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Create status chart
        chart1 = PieChart()
        chart1.title = "Tasks by Status"
        data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
        labels = Reference(ws, min_col=1, min_row=2, max_row=row-1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(labels)
        ws.add_chart(chart1, "D2")

        # Add priority chart data
        start_row = row + 2
        ws.cell(row=start_row, column=1, value='Priority')
        ws.cell(row=start_row, column=2, value='Count')
        row = start_row + 1

        for priority, count in priority_counts.items():
            ws.cell(row=row, column=1, value=priority)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Create priority chart
        chart2 = BarChart()
        chart2.title = "Tasks by Priority"
        chart2.style = 10
        data = Reference(ws, min_col=2, min_row=start_row, max_row=row-1)
        labels = Reference(ws, min_col=1, min_row=start_row+1, max_row=row-1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(labels)
        ws.add_chart(chart2, "D20")

    def _calculate_task_summary(self, tasks: List[Task]) -> Dict[str, Any]:
        """Calculate task summary statistics"""
        total = len(tasks)

        if total == 0:
            return {'Total Tasks': 0}

        # Status counts
        status_counts = {}
        for task in tasks:
            status_counts[task.status] = status_counts.get(task.status, 0) + 1

        # Priority counts
        priority_counts = {}
        for task in tasks:
            priority_counts[task.priority] = priority_counts.get(task.priority, 0) + 1

        # Calculate metrics
        completed = len([t for t in tasks if t.status in ['Resolved', 'Closed']])
        overdue = len([t for t in tasks if t.is_overdue()])

        # Average days open
        days_open_list = [t.days_open for t in tasks if t.days_open is not None]
        avg_days_open = sum(days_open_list) / len(days_open_list) if days_open_list else 0

        summary = {
            'Total Tasks': total,
            'Completed': completed,
            'Completion Rate': f"{(completed/total*100):.1f}%",
            'Overdue': overdue,
            'Average Days Open': f"{avg_days_open:.1f}"
        }

        # Add status breakdown
        for status, count in sorted(status_counts.items()):
            summary[f'{status} Tasks'] = count

        # Add priority breakdown
        for priority, count in sorted(priority_counts.items()):
            summary[f'{priority} Priority'] = count

        return summary

    def _create_summary_dataframe(self, tasks: List[Task]) -> pd.DataFrame:
        """Create summary DataFrame for tasks"""
        summary = self._calculate_task_summary(tasks)

        data = []
        for key, value in summary.items():
            data.append({'Metric': key, 'Value': value})

        return pd.DataFrame(data)

    def _create_pdf_summary(self, tasks: List[Task]) -> List:
        """Create PDF summary content"""
        if not REPORTLAB_AVAILABLE:
            return []

        story = []

        # Summary section
        story.append(Paragraph("Executive Summary", self.report_generator.styles['Heading1']))
        story.append(Spacer(1, 0.2 * inch))

        # Calculate summary
        summary = self._calculate_task_summary(tasks)

        # Create summary table
        summary_data = [['Metric', 'Value']]
        for key, value in summary.items():
            summary_data.append([key, str(value)])

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))

        return story

    def _create_pdf_task_table(self, tasks: List[Task]) -> Table:
        """Create PDF table for tasks"""
        if not REPORTLAB_AVAILABLE:
            return None

        # Prepare data
        data = [['Key', 'Title', 'Status', 'Priority', 'Due Date']]

        for task in tasks[:50]:  # Limit to 50 tasks for PDF
            data.append([
                task.key,
                task.title[:30] + '...' if len(task.title) > 30 else task.title,
                task.status,
                task.priority,
                task.target_date or 'N/A'
            ])

        # Create table
        table = Table(data, colWidths=[1*inch, 2.5*inch, 1.2*inch, 1*inch, 1*inch])

        # Style table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))

        return table

    def _add_department_summary(self, ws: Worksheet, team_members: List[TeamMember]):
        """Add department summary to worksheet"""
        if not OPENPYXL_AVAILABLE:
            return

        # Calculate department statistics
        dept_stats = {}
        for member in team_members:
            dept = member.department
            if dept not in dept_stats:
                dept_stats[dept] = {'total': 0, 'active': 0, 'roles': {}}

            dept_stats[dept]['total'] += 1
            if member.active:
                dept_stats[dept]['active'] += 1

            role = member.role
            dept_stats[dept]['roles'][role] = dept_stats[dept]['roles'].get(role, 0) + 1

        # Add to worksheet
        ws['A1'] = 'Department Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Headers
        headers = ['Department', 'Total Members', 'Active Members', 'Roles']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Data
        row = 4
        for dept, stats in sorted(dept_stats.items()):
            ws.cell(row=row, column=1, value=dept)
            ws.cell(row=row, column=2, value=stats['total'])
            ws.cell(row=row, column=3, value=stats['active'])

            # Format roles
            roles_text = ', '.join([f"{role}: {count}" for role, count in stats['roles'].items()])
            ws.cell(row=row, column=4, value=roles_text)

            row += 1

    def _add_category_summary(self, ws: Worksheet, legislation: List[LegislationReference]):
        """Add category summary to worksheet"""
        if not OPENPYXL_AVAILABLE:
            return

        # Calculate category statistics
        cat_stats = {}
        for leg in legislation:
            cat = leg.category
            if cat not in cat_stats:
                cat_stats[cat] = {'count': 0, 'jurisdictions': set()}

            cat_stats[cat]['count'] += 1
            cat_stats[cat]['jurisdictions'].add(leg.jurisdiction)

        # Add to worksheet
        ws['A1'] = 'Category Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')

        # Headers
        headers = ['Category', 'Count', 'Jurisdictions']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Data
        row = 4
        for cat, stats in sorted(cat_stats.items()):
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=stats['count'])
            ws.cell(row=row, column=3, value=', '.join(sorted(stats['jurisdictions'])))
            row += 1

    def _generate_excel_compliance_report(self, report_data: Dict[str, Any],
                                        filepath: Path) -> str:
        """Generate Excel compliance report"""
        if OPENPYXL_AVAILABLE:
            wb = Workbook()

            # Overview sheet
            ws_overview = wb.active
            ws_overview.title = "Overview"
            self._add_report_overview(ws_overview, report_data)

            # Tasks sheet
            if 'tasks' in report_data:
                ws_tasks = wb.create_sheet("Tasks")
                df_tasks = self._tasks_to_dataframe(report_data['tasks'])
                self._add_dataframe_to_worksheet(ws_tasks, df_tasks)

            # Team sheet
            if 'team' in report_data:
                ws_team = wb.create_sheet("Team")
                df_team = self._team_to_dataframe(report_data['team'])
                self._add_dataframe_to_worksheet(ws_team, df_team)

            # Compliance metrics sheet
            if 'metrics' in report_data:
                ws_metrics = wb.create_sheet("Metrics")
                self._add_compliance_metrics(ws_metrics, report_data['metrics'])

            wb.save(filepath)
        else:
            # Fallback to pandas
            with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
                # Overview
                overview_df = pd.DataFrame([{'Report Type': 'Compliance Report',
                                           'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')}])
                overview_df.to_excel(writer, sheet_name='Overview', index=False)

                # Tasks
                if 'tasks' in report_data:
                    df_tasks = self._tasks_to_dataframe(report_data['tasks'])
                    df_tasks.to_excel(writer, sheet_name='Tasks', index=False)

                # Team
                if 'team' in report_data:
                    df_team = self._team_to_dataframe(report_data['team'])
                    df_team.to_excel(writer, sheet_name='Team', index=False)

        return str(filepath)

    def _generate_pdf_compliance_report(self, report_data: Dict[str, Any],
                                      filepath: Path) -> str:
        """Generate PDF compliance report"""
        if not REPORTLAB_AVAILABLE:
            raise ValueError("ReportLab not available - cannot generate PDF")

        doc = SimpleDocTemplate(str(filepath), pagesize=letter)
        story = []

        # Title page
        story.append(Paragraph("Compliance Report", self.report_generator.styles['CustomTitle']))
        story.append(Spacer(1, 0.5 * inch))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            self.report_generator.styles['Normal']
        ))
        story.append(PageBreak())

        # Executive summary
        if 'summary' in report_data:
            story.append(Paragraph("Executive Summary", self.report_generator.styles['Heading1']))
            story.append(Spacer(1, 0.2 * inch))

            for key, value in report_data['summary'].items():
                story.append(Paragraph(f"<b>{key}:</b> {value}", self.report_generator.styles['Normal']))

            story.append(PageBreak())

        # Task summary
        if 'tasks' in report_data:
            story.extend(self._create_pdf_summary(report_data['tasks']))
            story.append(PageBreak())

        # Build PDF
        doc.build(story)
        return str(filepath)

    def _generate_json_compliance_report(self, report_data: Dict[str, Any],
                                       filepath: Path) -> str:
        """Generate JSON compliance report"""
        # Convert all objects to dictionaries
        export_data = {}

        for key, value in report_data.items():
            if key == 'tasks':
                export_data[key] = [task.to_dict() for task in value]
            elif key == 'team':
                export_data[key] = [member.to_dict() for member in value]
            elif key == 'legislation':
                export_data[key] = [leg.to_dict() for leg in value]
            else:
                export_data[key] = value

        # Add metadata
        export_data['metadata'] = {
            'generated': datetime.now().isoformat(),
            'version': self.config.version,
            'type': 'compliance_report'
        }

        # Write to file
        with open(filepath, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)

        return str(filepath)

    def _add_report_overview(self, ws: Worksheet, report_data: Dict[str, Any]):
        """Add report overview to worksheet"""
        if not OPENPYXL_AVAILABLE:
            return

        # Title
        ws['A1'] = 'Compliance Report Overview'
        ws['A1'].font = Font(size=18, bold=True)
        ws.merge_cells('A1:D1')

        # Report info
        ws['A3'] = 'Generated:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        ws['A4'] = 'Report Period:'
        ws['B4'] = report_data.get('period', 'All Time')

        # Summary statistics
        row = 6
        ws.cell(row=row, column=1, value='Summary Statistics')
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)

        row += 2
        if 'summary' in report_data:
            for key, value in report_data['summary'].items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1

    def _add_compliance_metrics(self, ws: Worksheet, metrics: Dict[str, Any]):
        """Add compliance metrics to worksheet"""
        if not OPENPYXL_AVAILABLE:
            return

        # Title
        ws['A1'] = 'Compliance Metrics'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')

        # Metrics data
        row = 3
        for category, category_metrics in metrics.items():
            # Category header
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1

            # Metrics
            if isinstance(category_metrics, dict):
                for metric, value in category_metrics.items():
                    ws.cell(row=row, column=2, value=metric)
                    ws.cell(row=row, column=3, value=value)
                    row += 1
            else:
                ws.cell(row=row, column=2, value=str(category_metrics))
                row += 1

            row += 1  # Add spacing between categories


# Global export service instance
_export_service = None


def get_export_service() -> ExportService:
    """Get global export service instance"""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service